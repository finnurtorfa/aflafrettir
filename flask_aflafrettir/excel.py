#!/usr/bin/python
# -*- coding: utf-8 -*-

""" 
excel.py
~~~~~~~~~~~

Contains functions for sorting and writing information contained in :class
Landings: to an excel file.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Color

style0 = Font(name='Arial', bold=False)
style1 = Font(name='Arial', bold=True)

item = [u'Skipaskrárnúmer', u'Nafn', u'Fjöldi', u'Heildarafli', u'Mesti afli',
    u'Höfn', u'Veiðarfæri', u'Afli e. tegundum']
keys = [u'shipNumber', u'shipName', u'count', u'totalCatch', u'maxCatch',
    u'landingHarbour', u'equipment', u'landingCatch']

def save_excel(name, landing_dict):
  """ Sets up a :class Workbook: and saves data in an excel file.

  :param name:          A string representing the name of the excel file.
  :param landing_dict:  A dictionary of landings to be written to excel file.
  """
  wbk = Workbook()

  for k in landing_dict:
    landing_dict[k].sort(reverse=True)
    write_data(wbk, k, landing_dict[k])
    
  wbk.remove_sheet(wbk.get_sheet_by_name('Sheet'))
  wbk.save(name)

def write_data(wbk, sheet_name, data_sorted):
  """ Writes data to a :class Workbook: object.

  :param wbk:         A :class Workbook: that will receive data
  :param sheet_name:  Name of the sheet to be added to the workbook. The sheet
                      name represents the group of landings in data_sorted.
  :param data_sorted: A list of :class Landings: sorted in reverse order by
                      totalCatch
  """
  sheet = wbk.create_sheet(title=sheet_name)

  for i, v in enumerate(item):
    c = sheet.cell(row=1, column=i+1, value=v)
    c.font = style1

  for i, v in enumerate(data_sorted):
    for j, k in enumerate(keys):
      tmp = getattr(v, k)

      if k == 'landingHarbour' or k == 'equipment':
        if type(tmp) is list:
          tmp = ', '.join([x for x in tmp])

      if k == 'landingCatch':
        for l, m in enumerate(tmp):
          c = sheet.cell(row=i+2,
                         column=j+1+l,
                         value=' : '.join([m, str(int(tmp[m]))]))
          c.font = style0
      else:
        if k == 'totalCatch' or k == 'maxCatch':
          tmp = tmp/1000.0

        c = sheet.cell(row=i+2,
                       column=j+1,
                       value=tmp)
        c.font = style0

__author__      = 'Finnur Smári Torfason'
__copyright__   = 'Copyright 2015, www.aflafrettir.is'
__credits__     = ['Finnur Smári Torfason', 'Gísli Reynisson']

__license__     = 'GPL v3'
__version__     = '0.2'
__maintainer__  = 'Finnur Smári Torfason'
__email__       = 'finnurtorfa@gmail.com'
__status__      = 'Development'
